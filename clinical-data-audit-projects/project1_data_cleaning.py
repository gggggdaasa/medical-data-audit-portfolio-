# clinical-data-audit-projects/project1_data_cleaning.py

import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def clean_clinical_data(input_path: str, output_path: str):
    """
    Cleans medical sheet data, standardizes dates to ISO 8601,
    flags duplicates, and outputs a stylized executive auditing sheet.
    """
    print("🚀 Initializing Clinical Data Clean & Auditing Pipeline...")
    
    # Generate mock dirty data if input file doesn't exist
    if not os.path.exists(input_path):
        print("⚠️ Input file not found. Generating mock dirty patient records...")
        dirty_data = {
            'PatientID': ['P-101', 'P-102', 'P-101', 'P-104', 'P-105', 'P-101', 'P-106'],
            'FullName': ['Ahmed Fawzi', 'Sarah Omar', 'Ahmed Fawzi', 'John Doe', 'Fatma Ali', 'Ahmd Fawzi', 'Mariam Hassan'],
            'DateOfRegistration': ['15-12-2025', '2025/11/20', '15-12-2025', '2026.01.12', '03-04-2026', '12/15/2025', 'abc-123-xyz'],
            'SysBP': [120, 145, 120, 999, 115, 120, 130],  # 999 is an obvious medical outlier
            'DiaBP': [80, 92, 80, 60, 75, 80, None]
        }
        df = pd.DataFrame(dirty_data)
        df.to_csv(input_path, index=False)
        print(f"✅ Mock raw clinical data saved to {input_path}")
    
    # 1. Load Data
    df = pd.read_csv(input_path)
    
    # 2. Advanced Deduplication
    print("🔍 Auditing duplicate ID records and patient logs...")
    # Tag true exact duplicates
    df['Audit_IsDuplicate'] = df.duplicated(subset=['PatientID', 'FullName'], keep='first')
    
    # Clean fuzzy duplicates manually or based on PatientID
    df_cleaned = df[~df['Audit_IsDuplicate']].copy()
    
    # 3. Date Standardization (to ISO 8601 format: YYYY-MM-DD)
    print("📅 Standardizing clinical registration timestamps to ISO 8601...")
    def standardize_date(date_str):
        if pd.isna(date_str):
            return "N/A"
        date_str = str(date_str).strip()
        for fmt in ('%d-%m-%Y', '%Y/%m/%d', '%Y.%m.%d', '%m/%d/%Y'):
            try:
                return pd.to_datetime(date_str, format=fmt).strftime('%Y-%m-%d')
            except ValueError:
                continue
        # Fallback using dateutil or default pandas parser
        try:
            return pd.to_datetime(date_str).strftime('%Y-%m-%d')
        except Exception:
            return "INVALID_FORMAT"

    df_cleaned['DateOfRegistration'] = df_cleaned['DateOfRegistration'].apply(standardize_date)
    
    # 4. Outlier Flagging (e.g. SysBP clinical ranges between 60 and 250)
    print("🩺 Applying physiological range threshold filters...")
    df_cleaned['BP_Audit_Status'] = np.where(
        (df_cleaned['SysBP'] < 70) | (df_cleaned['SysBP'] > 220) | 
        (df_cleaned['DiaBP'].isna()) | (df_cleaned['DiaBP'] < 40) | (df_cleaned['DiaBP'] > 130),
        'OUTLIER / INVALID', 'VALID'
    )
    
    # Impute normal placeholder values or flag outliers
    df_cleaned.loc[df_cleaned['SysBP'] == 999, 'SysBP'] = 120 # Logical clinical correction
    
    # 5. Styling & Writing to Excel with openpyxl
    print("✏️ Structuring and styling output audit report using openpyxl...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Patient Registration Audit"
    ws.views.sheetView[0].showGridLines = True
    
    # Formulate Header font/row styling
    header_fill = PatternFill(start_color="1e293b", end_color="0f172a", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="ffffff")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Write styled columns headers
    cols = list(df_cleaned.columns)
    ws.append(cols)
    for col_idx in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Add cleaned dataset rows
    border_style = Side(border_style="thin", color="cbd5e1")
    cell_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    data_font = Font(name="Segoe UI", size=10)
    
    for r in dataframe_to_rows(df_cleaned, index=False, header=False):
        ws.append(r)
        
    # Apply soft highlights for medical outliers and grid alignment
    error_fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid") # soft red
    error_font = Font(name="Segoe UI", size=10, color="991b1b", bold=True)
    
    for row in range(2, ws.max_row + 1):
        bp_status_cell = ws.cell(row=row, column=cols.index('BP_Audit_Status') + 1)
        # Apply error styling on clinical outlier rows
        if bp_status_cell.value == "OUTLIER / INVALID":
            for col in range(1, len(cols) + 1):
                c = ws.cell(row=row, column=col)
                c.fill = error_fill
                c.font = error_font
        else:
            for col in range(1, len(cols) + 1):
                c = ws.cell(row=row, column=col)
                c.font = data_font
                
        # Thin Slate Border configuration
        for col in range(1, len(cols) + 1):
            ws.cell(row=row, column=col).border = cell_border

    # Standardize column auto-widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    wb.save(output_path)
    print(f"🎉 Pipeline Execution Complete! Cleaned and audited report saved to:\n➡️ {output_path}")

if __name__ == "__main__":
    clean_clinical_data("raw_patient_records.csv", "Cleaned_Patient_Medical_Audit.xlsx")
